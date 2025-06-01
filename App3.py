import os
from flask import Flask, render_template, request, redirect, session, flash, url_for
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
from flask import make_response
from xhtml2pdf import pisa
from io import BytesIO
from flask import Response
from datetime import datetime

import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

app = Flask(__name__)

# === Configuration ===
app.secret_key = 'your_secret_key_here'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///tasks.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Image Upload Paths
# BASE_UPLOAD_FOLDER = r'C:\Users\VENKATESH\OneDrive\Desktop\KSR_Legal\static\images'
# CA_IMAGE_FOLDER = os.path.join(BASE_UPLOAD_FOLDER, 'ca')
# os.makedirs(CA_IMAGE_FOLDER, exist_ok=True)

BASE_UPLOAD_FOLDER = r'C:\Users\VENKATESH\OneDrive\Desktop\KSR_Legal\static\images'
app.config['CA_IMAGE_FOLDER'] = os.path.join(BASE_UPLOAD_FOLDER, 'ca')
app.config['BETTA_IMAGE_FOLDER'] = os.path.join(BASE_UPLOAD_FOLDER, 'betta')
app.config['PETITIONS_IMAGE_FOLDER'] = os.path.join(BASE_UPLOAD_FOLDER, 'petitions')
app.config['NEWCASE_IMAGE_FOLDER'] = os.path.join(BASE_UPLOAD_FOLDER, 'new_cases')

os.makedirs(app.config['CA_IMAGE_FOLDER'], exist_ok=True)
os.makedirs(app.config['BETTA_IMAGE_FOLDER'], exist_ok=True)
os.makedirs(app.config['PETITIONS_IMAGE_FOLDER'], exist_ok=True)
os.makedirs(app.config['NEWCASE_IMAGE_FOLDER'], exist_ok=True)


# app.config['UPLOAD_FOLDER'] = CA_IMAGE_FOLDER
app.config['ALLOWED_EXTENSIONS'] = {'png', 'jpg', 'jpeg', 'gif'}

# === Initialization ===
db = SQLAlchemy(app)

# === Helpers ===
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        user = User.query.get(session.get('user_id'))
        if not user or user.username != 'admin':
            flash("Admin access required.")
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapper

# === Models ===
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(150), unique=True, nullable=False)
    password_hash = db.Column(db.String(150), nullable=False)

class Task(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    case_number = db.Column(db.String(100))
    ca_filed_date = db.Column(db.String(100))
    ca_number = db.Column(db.String(100))
    required_details = db.Column(db.Text)
    filed_for = db.Column(db.String(100))
    cf_date = db.Column(db.String(100))
    cf_amount = db.Column(db.String(100))
    cf_paid_date = db.Column(db.String(100))
    ca_ready_date = db.Column(db.String(100))
    remarks = db.Column(db.Text)
    status = db.Column(db.String(50))
    image1 = db.Column(db.String(255))
    image2 = db.Column(db.String(255))

class BettaFiling(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    court_name = db.Column(db.String(100))
    case_no = db.Column(db.String(100))
    betta_paid_on = db.Column(db.String(100))
    betta_paid_for = db.Column(db.String(100))
    gr_number = db.Column(db.String(100))
    gr_date = db.Column(db.String(100))
    betta_returned_date = db.Column(db.String(100))
    psa_numbers = db.Column(db.String(100))
    status = db.Column(db.String(50))
    remarks = db.Column(db.String(255))
    image1 = db.Column(db.String(255))
    image2 = db.Column(db.String(255))

class PetitionFiling(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    court_name = db.Column(db.String(100))
    case_no = db.Column(db.String(100))
    petition_filed_date = db.Column(db.String(100))
    petition_filed_for = db.Column(db.String(100))
    gr_number = db.Column(db.String(100))
    gr_date = db.Column(db.String(100))
    hearing_date = db.Column(db.String(100))
    status = db.Column(db.String(50))
    image1 = db.Column(db.String(255))
    image2 = db.Column(db.String(255))

class NewCaseFiling(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    court_name = db.Column(db.String(100))
    filed_date = db.Column(db.String(100))
    nature_of_case = db.Column(db.String(100))
    case_no = db.Column(db.String(100))
    gr_number = db.Column(db.String(100))
    gr_date = db.Column(db.String(100))
    plaintiff_name = db.Column(db.String(100))
    defendant_name = db.Column(db.String(100))
    made_over_court = db.Column(db.String(100))
    hearing_date = db.Column(db.String(100))
    status = db.Column(db.String(50))
    image1 = db.Column(db.String(255))
    image2 = db.Column(db.String(255))

# === Routes ===


@app.route('/')
@admin_required
def index():
    tasks = Task.query.all()
    bettas = BettaFiling.query.all()
    petitions = PetitionFiling.query.all()
    new_cases = NewCaseFiling.query.all()
    return render_template('dashboard.html', tasks=tasks, bettas=bettas, petitions=petitions, new_cases=new_cases)

@app.route('/dashboard')
@admin_required
def dashboard():
    tasks = Task.query.all()
    bettas = BettaFiling.query.all()
    petitions = PetitionFiling.query.all()
    new_cases = NewCaseFiling.query.all()
    return render_template('dashboard.html', tasks=tasks, bettas=bettas, petitions=petitions, new_cases=new_cases)


@app.route('/add', methods=['GET', 'POST'])
@admin_required
def add():
    if request.method == 'POST':
        task = Task(**request.form)

        image1 = request.files.get('image1')
        if image1 and allowed_file(image1.filename):
            filename1 = secure_filename(image1.filename)
            image1.save(os.path.join(app.config['CA_IMAGE_FOLDER'], filename1))
            task.image1 = filename1

        image2 = request.files.get('image2')
        if image2 and allowed_file(image2.filename):
            filename2 = secure_filename(image2.filename)
            image2.save(os.path.join(app.config['CA_IMAGE_FOLDER'], filename2))
            task.image2 = filename2

        db.session.add(task)
        db.session.commit()
        return redirect('/dashboard')
    return render_template('form.html', task=None)



@app.route('/update/<int:task_id>', methods=['GET', 'POST'])
@admin_required
def update(task_id):
    task = Task.query.get_or_404(task_id)
    if request.method == 'POST':
        for field in request.form:
            setattr(task, field, request.form[field])
        # Handle updated images
        image1 = request.files.get('image1')
        if image1 and allowed_file(image1.filename):
            filename1 = secure_filename(image1.filename)
            image1.save(os.path.join(app.config['CA_IMAGE_FOLDER'], filename1))
            task.image1 = filename1

        image2 = request.files.get('image2')
        if image2 and allowed_file(image2.filename):
            filename2 = secure_filename(image2.filename)
            image2.save(os.path.join(app.config['CA_IMAGE_FOLDER'], filename2))
            task.image2 = filename2
        db.session.commit()
        return redirect('/dashboard')
    return render_template('form.html', task=task)

@app.route('/delete/<int:task_id>')
@admin_required
def delete(task_id):
    task = Task.query.get_or_404(task_id)
    db.session.delete(task)
    db.session.commit()
    return redirect('/dashboard')

@app.route('/betta/delete/<int:id>')
@admin_required
def delete_new_betta(id):
    entry = BettaFiling.query.get_or_404(id)
    db.session.delete(entry)
    db.session.commit()
    return redirect('/dashboard')


@app.route('/petition/delete/<int:id>')
@admin_required
def delete_petition(id):
    entry = PetitionFiling.query.get_or_404(id)
    db.session.delete(entry)
    db.session.commit()
    return redirect('/dashboard')

@app.route('/newcase/delete/<int:id>')
@admin_required
def delete_new_case(id):
    entry = NewCaseFiling.query.get_or_404(id)
    db.session.delete(entry)
    db.session.commit()
    return redirect('/dashboard')


## Betta Dashboard ##

@app.route('/betta/add', methods=['GET', 'POST'])
@admin_required
def add_betta():
    if request.method == 'POST':
        data = request.form
        new_entry = BettaFiling(**data)

        img1 = request.files.get('image1')
        img2 = request.files.get('image2')

        if img1 and allowed_file(img1.filename):
            fname = secure_filename(img1.filename)
            img1.save(os.path.join(app.config['BETTA_IMAGE_FOLDER'], fname))
            new_entry.image1 = fname
        if img2 and allowed_file(img2.filename):
            fname = secure_filename(img2.filename)
            img2.save(os.path.join(app.config['BETTA_IMAGE_FOLDER'], fname))
            new_entry.image2 = fname

        db.session.add(new_entry)
        db.session.commit()
        return redirect('/dashboard')
    return render_template('betta_form.html', entry=None)

@app.route('/betta/update/<int:id>', methods=['GET', 'POST'])
@admin_required
def update_betta(id):
    entry = BettaFiling.query.get_or_404(id)
    if request.method == 'POST':
        for key in request.form:
            setattr(entry, key, request.form[key])

        img1 = request.files.get('image1')
        img2 = request.files.get('image2')

        if img1 and allowed_file(img1.filename):
            fname = secure_filename(img1.filename)
            img1.save(os.path.join(app.config['BETTA_IMAGE_FOLDER'], fname))
            entry.image1 = fname
        if img2 and allowed_file(img2.filename):
            fname = secure_filename(img2.filename)
            img2.save(os.path.join(app.config['BETTA_IMAGE_FOLDER'], fname))
            entry.image2 = fname

        db.session.commit()
        return redirect('/dashboard')
    return render_template('betta_form.html', entry=entry)

##Petitions Dashboard##

@app.route('/petition/add', methods=['GET', 'POST'])
@admin_required
def add_petition():
    if request.method == 'POST':
        entry = PetitionFiling(**request.form)

        img1 = request.files.get('image1')
        img2 = request.files.get('image2')

        if img1 and allowed_file(img1.filename):
            fname = secure_filename(img1.filename)
            img1.save(os.path.join(app.config['PETITIONS_IMAGE_FOLDER'], fname))
            entry.image1 = fname

        if img2 and allowed_file(img2.filename):
            fname = secure_filename(img2.filename)
            img2.save(os.path.join(app.config['PETITIONS_IMAGE_FOLDER'], fname))
            entry.image2 = fname

        db.session.add(entry)
        db.session.commit()
        return redirect('/dashboard')
    return render_template('petition_form.html', entry=None)

@app.route('/petition/update/<int:id>', methods=['GET', 'POST'])
@admin_required
def update_petition(id):
    entry = PetitionFiling.query.get_or_404(id)
    if request.method == 'POST':
        for field in request.form:
            setattr(entry, field, request.form[field])

        img1 = request.files.get('image1')
        img2 = request.files.get('image2')

        if img1 and allowed_file(img1.filename):
            fname = secure_filename(img1.filename)
            img1.save(os.path.join(app.config['PETITIONS_IMAGE_FOLDER'], fname))
            entry.image1 = fname

        if img2 and allowed_file(img2.filename):
            fname = secure_filename(img2.filename)
            img2.save(os.path.join(app.config['PETITIONS_IMAGE_FOLDER'], fname))
            entry.image2 = fname

        db.session.commit()
        return redirect('/dashboard')
    return render_template('petition_form.html', entry=entry)

##New Case Dashboard##

@app.route('/newcase/add', methods=['GET', 'POST'])
@admin_required
def add_new_case():
    if request.method == 'POST':
        entry = NewCaseFiling(**request.form)

        img1 = request.files.get('image1')
        img2 = request.files.get('image2')

        if img1 and allowed_file(img1.filename):
            fname = secure_filename(img1.filename)
            img1.save(os.path.join(app.config['NEWCASE_IMAGE_FOLDER'], fname))
            entry.image1 = fname

        if img2 and allowed_file(img2.filename):
            fname = secure_filename(img2.filename)
            img2.save(os.path.join(app.config['NEWCASE_IMAGE_FOLDER'], fname))
            entry.image2 = fname

        db.session.add(entry)
        db.session.commit()
        return redirect('/dashboard')
    return render_template('newcase_form.html', entry=None)

@app.route('/newcase/update/<int:id>', methods=['GET', 'POST'])
@admin_required
def update_new_case(id):
    entry = NewCaseFiling.query.get_or_404(id)
    if request.method == 'POST':
        for field in request.form:
            setattr(entry, field, request.form[field])

        img1 = request.files.get('image1')
        img2 = request.files.get('image2')

        if img1 and allowed_file(img1.filename):
            fname = secure_filename(img1.filename)
            img1.save(os.path.join(app.config['NEWCASE_IMAGE_FOLDER'], fname))
            entry.image1 = fname

        if img2 and allowed_file(img2.filename):
            fname = secure_filename(img2.filename)
            img2.save(os.path.join(app.config['NEWCASE_IMAGE_FOLDER'], fname))
            entry.image2 = fname

        db.session.commit()
        return redirect('/dashboard')
    return render_template('newcase_form.html', entry=entry)



@app.route('/download/pdf')
@admin_required
def download_pdf():
    tasks = Task.query.all()
    bettas = BettaFiling.query.all()
    petitions = PetitionFiling.query.all()
    new_cases = NewCaseFiling.query.all()

    html = render_template(
        'pdf_template.html',
        tasks=tasks,
        bettas=bettas,
        petitions=petitions,
        new_cases=new_cases,
        datetime_now=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    )

    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    
    if not pdf.err:
        response = make_response(result.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = 'attachment; filename=KSR_Filing_Dashboard.pdf'
        return response
    return "PDF generation failed"

@app.route('/download/excel')
@admin_required
def download_excel():
    # Load all data
    task_data = Task.query.all()
    betta_data = BettaFiling.query.all()
    petition_data = PetitionFiling.query.all()
    newcase_data = NewCaseFiling.query.all()

    wb = Workbook()
    wb.remove(wb.active)

    # === Helper: Add sheet from list of objects ===
    def add_sheet(sheet_name, queryset, columns):
        ws = wb.create_sheet(title=sheet_name)
        df = pd.DataFrame([{col: getattr(obj, col) for col in columns} for obj in queryset])

        if df.empty:
            ws.append(["No data available"])
            return

        # Add header
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # Apply color coding
        status_col_idx = columns.index('status') + 1  # Excel columns start at 1
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(columns)):
            status_cell = row[status_col_idx - 1]
            status = (status_cell.value or '').strip().lower()

            fill = None
            if 'in progress' in status:
                fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # Light Yellow
            elif 'completed' in status:
                fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # Light Green
            elif 'error' in status:
                fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")  # Light Red
            elif 'served' in status:
                fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # Light Green
            elif 'unserved' in status:
                fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")  # Light Red
            else:
                fill = PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid")  # Light Gray

            for cell in row:
                cell.fill = fill

    # Define fields for each model
    add_sheet("CA Dashboard", task_data, ['case_number', 'ca_filed_date', 'ca_number', 'filed_for', 'status'])
    add_sheet("BETTA Filing", betta_data, ['court_name', 'case_no', 'betta_paid_on', 'betta_paid_for', 'status'])
    add_sheet("PETITIONS Filing", petition_data, ['court_name', 'case_no', 'petition_filed_date', 'petition_filed_for', 'status'])
    add_sheet("NEW CASE Filing", newcase_data, ['court_name', 'filed_date', 'nature_of_case', 'case_no', 'status'])

    # Save to memory
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=KSR_Filing_Dashboard.xlsx"}
    )

# === Auth ===
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username'].strip()
        password = request.form['password'].strip()
        if not username or not password:
            flash("Username and password required.")
            return redirect('/register')
        if User.query.filter_by(username=username).first():
            flash("Username already exists.")
            return redirect('/register')
        hashed_pw = generate_password_hash(password)
        db.session.add(User(username=username, password_hash=hashed_pw))
        db.session.commit()
        flash("Registered successfully.")
        return redirect('/login')
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form['username']).first()
        if user and check_password_hash(user.password_hash, request.form['password']):
            session['user_id'] = user.id
            return redirect('/')
        flash("Invalid credentials.")
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash("Logged out.")
    return redirect('/login')

if __name__ == '__main__':
    db.create_all()
    app.run(debug=True)
